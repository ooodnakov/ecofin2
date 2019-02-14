import openpyxl
import wr
from openpyxl.utils import get_column_letter as let


def totable():
    wb = openpyxl.Workbook()
    ws = wb.active
    results = wr.read_results()
    ids = list(results.keys())
    al = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    for row in ws['A1:ZZ100']:
        for cell in row:
            cell.alignment = al
    ws.column_dimensions['A'].width = 18
    for i in range(2, 1000):
        ws.column_dimensions[let(i)].width = 5
    i = 0
    for id in ids:
        ws.cell(row=2 + 3 * i, column=1, value=results[ids[i]][1])
        ws.cell(row=3 + 3 * i, column=1, value=ids[i])
        i += 1
    index = 0
    for id in ids:
        for j in range(0, len(results[id][2])):
            a = ws.cell(
                row=1 + index * 3,
                column=2 + 13 * j,
                value=results[id][2][j][0])
            a = ws.merge_cells(
                start_row=1 + index * 3,
                start_column=2 + 13 * j,
                end_row=1 + index * 3,
                end_column=14 + 13 * j)
            cells = [{}, {}]
            for i in range(1, 13):
                cells[0][i] = ws.cell(row=2 + index * 3,
                                      column=1 + int(i) + j * 13,
                                      value=results[id][2][j][1][str(i)][0])
            for i in range(1, 13):
                try:
                    cells[1][i] = ws.cell(row=3 + index * 3,
                                          column=1 + int(i) + j * 13,
                                          value=results[id][2][j][1][str(i)][1])
                except IndexError:
                    cells[1][i] = ws.cell(
                        row=3 + index * 3, column=1 + int(i) + j * 13, value=0)
            ws.cell(
                row=2 + index * 3,
                column=14 + j * 13,
                value=results[id][2][j][2])
            ws.cell(
                row=3 + index * 3,
                column=14 + j * 13,
                value=results[id][2][j][3])

        index += 1

    wb.save('res.xlsx')
